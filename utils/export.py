# utils/export.py
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

HEADERS = ["Nom", "Type", "Téléphone", "Adresse", "Note", "Nb avis", "Google Maps"]
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generate_excel(leads: list[dict], output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row, lead in enumerate(leads, start=2):
        ws.cell(row=row, column=1, value=lead.get("name", ""))
        ws.cell(row=row, column=2, value=lead.get("type", ""))
        ws.cell(row=row, column=3, value=lead.get("phone", ""))
        ws.cell(row=row, column=4, value=lead.get("address", ""))
        ws.cell(row=row, column=5, value=lead.get("rating"))
        ws.cell(row=row, column=6, value=lead.get("user_ratings_total", 0))
        ws.cell(row=row, column=7, value=lead.get("maps_url", ""))

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    wb.save(output_path)
