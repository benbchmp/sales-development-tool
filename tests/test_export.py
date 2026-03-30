# tests/test_export.py
import os
from openpyxl import load_workbook
from utils.export import generate_excel

SAMPLE_LEADS = [
    {
        "name": "Dupont Plomberie", "type": "plumber", "phone": "0601020304",
        "address": "12 rue de la Paix, 69001 Lyon", "rating": 4.5,
        "user_ratings_total": 12, "maps_url": "https://maps.google.com/?cid=p1",
    },
    {
        "name": "Martin Électricité", "type": "electrician", "phone": "0607080910",
        "address": "5 avenue Foch, 69006 Lyon", "rating": 4.0,
        "user_ratings_total": 5, "maps_url": "https://maps.google.com/?cid=p2",
    },
]


def test_generate_excel_creates_file(tmp_path):
    output_path = str(tmp_path / "leads.xlsx")
    generate_excel(SAMPLE_LEADS, output_path)
    assert os.path.exists(output_path)


def test_generate_excel_has_correct_headers(tmp_path):
    output_path = str(tmp_path / "leads.xlsx")
    generate_excel(SAMPLE_LEADS, output_path)
    wb = load_workbook(output_path)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, 8)]
    assert headers == ["Nom", "Type", "Téléphone", "Adresse", "Note", "Nb avis", "Google Maps"]


def test_generate_excel_has_correct_data(tmp_path):
    output_path = str(tmp_path / "leads.xlsx")
    generate_excel(SAMPLE_LEADS, output_path)
    wb = load_workbook(output_path)
    ws = wb.active
    assert ws.cell(2, 1).value == "Dupont Plomberie"
    assert ws.cell(2, 3).value == "0601020304"
    assert ws.cell(3, 1).value == "Martin Électricité"
